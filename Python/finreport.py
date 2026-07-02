# Финотчёт ВБ: разрез по артикулам. Запуск: .venv/bin/python finreport.py → http://localhost:8010
# ponytail: всё в памяти процесса — перезапуск стирает данные; БД добавить, когда понадобится история.
import csv, io, sys
from collections import defaultdict
from flask import Flask, request, redirect, send_file
from openpyxl import load_workbook

app = Flask(__name__)

ART = defaultdict(lambda: defaultdict(float))  # nm_id -> метрики из детализации
OTHER = defaultdict(lambda: defaultdict(float))  # nm_id -> тип удержания -> сумма
COST = {}   # nm_id -> себестоимость за штуку
ADV = defaultdict(float)  # nm_id -> затраты на РК
TAX = {"mode": "", "choice": "", "custom": ""}  # налоговый режим; choice: 6/15/25/other/none
SKIP = {"promo": False}  # прятать «Оказание услуг WB Продвижение» — фильтр на показе, данные не трогает


def is_promo(label):
    return "wb продвижение" in label.lower()

INCOME_MODES = ("УСН «Доходы»", "НПД")  # база — вся выручка; иначе — прибыль до налога
TAX_DEFAULTS = {"УСН «Доходы»": "6", "УСН «Доходы минус расходы»": "15", "НПД": "other", "ОСН": "25"}


def tax_rate():
    c = TAX["choice"]
    if not c or c == "none":
        return None
    return (num(TAX["custom"]) or None) if c == "other" else float(c)

# префикс нормализованного заголовка -> ключ метрики
HEADERS = {
    "код номенклатуры": "nm",
    "артикул поставщика": "supplier_art",
    "тип документа": "doc",
    "обоснование для оплаты": "oplata",
    "кол-во": "qty",
    "вайлдберриз реализовал товар": "sold",
    "к перечислению продавцу": "pay",
    "услуги по доставке товара покупателю": "logistics",
    "общая сумма штрафов": "fines",
    "хранение": "storage",
    "операции при приёмке": "acceptance",
    "удержания": "deduction",
    "корректировка вознаграждения": "vv_corr",
    "возмещение издержек": "reimb",
    "стоимость участия в программе лояльности": "loy_cost",
    "сумма, удержанная за начисленные баллы": "loy_points",
    "компенсация скидки по программе лояльности": "loy_comp",
    "виды логистики": "vid",  # «Виды логистики, штрафов и корректировок ВВ» — расшифровка удержаний
}


def num(v):
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def parse_report(fp):
    ws = load_workbook(fp, read_only=True, data_only=True).active
    rows = ws.iter_rows(values_only=True)
    head = next(rows)
    col = {}
    for i, h in enumerate(head):
        hn = " ".join(str(h or "").lower().split())
        for prefix, key in HEADERS.items():
            if hn.startswith(prefix) and key not in col:
                col[key] = i
    if "nm" not in col or "pay" not in col:
        raise ValueError("не похоже на детализацию ВБ: нет столбцов «Код номенклатуры»/«К перечислению»")
    ART.clear()
    OTHER.clear()
    for r in rows:
        g = lambda k: r[col[k]] if k in col and col[k] < len(r) else None
        nm = str(g("nm") or "").strip()
        nm = nm if nm not in ("", "0", "None") else "— без артикула"
        a = ART[nm]
        if g("supplier_art"):
            a["supplier_art"] = g("supplier_art")
        oplata = str(g("oplata") or "").strip().lower()
        sign = -1 if str(g("doc") or "").strip().lower() == "возврат" else 1
        if oplata == "продажа":
            a["qty_sold"] += num(g("qty"))
        elif oplata == "возврат":
            a["qty_ret"] += num(g("qty"))
        a["sold"] += sign * num(g("sold"))
        a["pay"] += sign * num(g("pay"))
        a["logistics"] += num(g("logistics"))
        a["fines"] += num(g("fines"))
        a["storage"] += num(g("storage"))
        a["acceptance"] += num(g("acceptance"))
        # ponytail: всё редкое — в одну корзину «прочие удержания»; знаки по инструкции ВБ (удержание «+», выплата «−»)
        # тип берём из «Виды логистики, штрафов и корректировок ВВ», иначе из обоснования
        vid = str(g("vid") or "").strip()

        def oth(label, val):
            if val:
                a["other"] += val
                OTHER[nm][label] += val
        oth(vid or str(g("oplata") or "").strip() or "Удержания", num(g("deduction")))
        oth(vid or "Корректировка ВВ", num(g("vv_corr")))
        oth("Возмещение издержек", num(g("reimb")))
        oth("Участие в программе лояльности", num(g("loy_cost")))
        oth("Баллы программы лояльности", num(g("loy_points")))
        oth("Компенсация скидки ПЛ", -num(g("loy_comp")))


def rows_from_upload(f):
    """xlsx или csv → список строк-списков."""
    if f.filename.lower().endswith(".xlsx"):
        ws = load_workbook(f, read_only=True, data_only=True).active
        return [list(r) for r in ws.iter_rows(values_only=True)]
    text = f.read().decode("utf-8-sig", errors="replace")
    delim = ";" if ";" in text.splitlines()[0] else ("\t" if "\t" in text.splitlines()[0] else ",")
    return list(csv.reader(io.StringIO(text), delimiter=delim))


def fmt(v):
    return f"{v:,.0f}".replace(",", " ") if v else ""


@app.post("/report")
def upload_report():
    parse_report(request.files["f"])
    return redirect("/")


@app.post("/skip")
def set_skip():
    SKIP["promo"] = bool(request.form.get("on"))
    return redirect("/")


@app.post("/cost")
def upload_cost():
    for r in rows_from_upload(request.files["f"]):
        if len(r) >= 2 and num(r[1]):
            COST[str(r[0]).strip()] = num(r[1])
    return redirect("/")


@app.post("/tax")
def set_tax():
    TAX.update(mode=request.form.get("mode", ""), choice=request.form.get("rate", "none"),
               custom=request.form.get("custom", ""))
    return redirect("/")


@app.post("/adv")
def upload_adv():
    ADV.clear()
    for r in rows_from_upload(request.files["f"]):
        if len(r) >= 3 and num(r[2]):
            ADV[str(r[1]).strip()] += num(r[2])
    return redirect("/")


@app.get("/export")
def export():
    from openpyxl import Workbook
    cols, rows, foot, rate = compute(request.args.get("other"), request.args.get("tax"))
    wb = Workbook()
    ws = wb.active
    ws.title = "По артикулам"

    def x(c, v):
        if v is None:
            return ""
        if c.startswith("↳ Ставка") or c == "Маржа, %":
            return f"{round(v, 1):g}".replace(".", ",")  # десятичный разделитель — запятая
        return round(v)
    ws.append(cols)
    for row in [foot] + rows:
        ws.append([row[0], row[1]] + [x(c, v) for c, v in zip(cols[2:], row[2:])])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name="finreport.xlsx", as_attachment=True)


def compute(o_flag, t_flag):
    """Собранная таблица: (cols, rows, foot, rate). Значения сырые числа; маржа/итог могут быть None."""
    rate = tax_rate()
    skip = SKIP["promo"]
    type_tot = defaultdict(float)
    for d in OTHER.values():
        for t, v in d.items():
            if not (skip and is_promo(t)):
                type_tot[t] += v
    types = sorted(type_tot, key=lambda t: -abs(type_tot[t])) if o_flag else []
    cols = ["Артикул ВБ", "Артикул продавца", "Продаж, шт", "Возвратов, шт", "ВБ реализовал",
            "К перечислению", "Логистика", "Хранение", "Приёмка", "Штрафы",
            "Прочие удержания"] + [f"↳ {t}" for t in types] + ["Реклама", "Себестоимость"]
    if rate is not None:
        cols += ["Налог"] + (["↳ Налоговая база", "↳ Ставка, %"] if t_flag else [])
    cols += ["Прибыль", "Маржа, %"]

    rows, total = [], defaultdict(float)
    for nm in sorted(ART, key=lambda n: -ART[n]["pay"]):
        a = ART[nm]
        cogs = COST.get(nm, 0) * (a["qty_sold"] - a["qty_ret"])
        adv = ADV.get(nm, 0)
        other = a["other"] - (sum(v for t, v in OTHER[nm].items() if is_promo(t)) if skip else 0)
        pre = (a["pay"] - a["logistics"] - a["storage"] - a["acceptance"]
               - a["fines"] - other - cogs - adv)
        # ponytail: налоговая база упрощённо — выручка (УСН Доходы/НПД) либо прибыль до налога (Д−Р/ОСН), без НДС и нюансов
        base = max(a["sold"] if TAX["mode"] in INCOME_MODES else pre, 0) if rate is not None else 0
        tax = base * (rate or 0) / 100
        profit = pre - tax
        vals = [a["qty_sold"], a["qty_ret"], a["sold"], a["pay"], a["logistics"], a["storage"],
                a["acceptance"], a["fines"], other] + \
               [OTHER[nm].get(t, 0) for t in types] + [adv, cogs]
        if rate is not None:
            vals += [tax] + ([base, rate] if t_flag else [])
        vals += [profit]
        for c, v in zip(cols[2:-1], vals):
            total[c] += v
        rows.append([nm, a.get("supplier_art", "")] + vals +
                    [profit / a["sold"] * 100 if a["sold"] else None])
    t_margin = total["Прибыль"] / total["ВБ реализовал"] * 100 if total["ВБ реализовал"] else None
    foot = ["ИТОГО", ""] + [rate if c.startswith("↳ Ставка") else total[c] for c in cols[2:-1]] + [t_margin]
    return cols, rows, foot, rate


@app.get("/")
def index():
    o_flag, t_flag = request.args.get("other"), request.args.get("tax")

    def link(o, t, text, path="/"):
        p = (["other=1"] if o else []) + (["tax=1"] if t else [])
        return f'<a href="{path}{"?" + "&".join(p) if p else ""}">{text}</a>'
    cols, rows, foot_vals, rate = compute(o_flag, t_flag)
    skip = SKIP["promo"]

    def tr(row, cls=""):
        cells = [row[0], row[1]] + \
                ["" if v is None else f"{v:g}" if c.startswith("↳ Ставка") else fmt(v)
                 for c, v in zip(cols[2:-1], row[2:-1])] + \
                ["" if row[-1] is None else f"{row[-1]:.1f}"]
        return f"<tr{cls}>" + "".join(f"<td>{c}</td>" for c in cells) + "</tr>"
    body = "".join(tr(r) for r in rows)
    foot = tr(foot_vals, " class=tot")
    hdr = {"Прочие удержания": "Прочие удержания " + link(not o_flag, t_flag, "[свернуть]" if o_flag else "[раскрыть]"),
           "Налог": "Налог " + link(o_flag, not t_flag, "[свернуть]" if t_flag else "[раскрыть]")}
    head_html = "".join(f"<th>{hdr.get(c, c)}</th>" for c in cols)
    export_link = link(o_flag, t_flag, "⬇ Скачать XLSX", "/export")

    # модалка налогов
    sel_mode = TAX["mode"] or "УСН «Доходы»"
    choice = TAX["choice"] or "6"
    mode_opts = "".join(f'<option {"selected" if m == sel_mode else ""}>{m}</option>' for m in TAX_DEFAULTS)
    radios = ""
    for v, txt in [("6", "6%"), ("15", "15%"), ("25", "25%"), ("other", "Другая:"), ("none", "Не учитывать")]:
        radios += f'<label><input type=radio name=rate value={v} {"checked" if v == choice else ""}> {txt}</label>'
        radios += f' <input name=custom size=6 value="{TAX["custom"]}"><br>' if v == "other" else "<br>"
    tax_now = f"Сейчас: {TAX['mode']}, {f'{rate:g}%' if rate is not None else 'не учитывается'}" if TAX["mode"] else ""
    modal = f"""<button type=button onclick="document.getElementById('taxdlg').showModal()">Настроить налоги</button> <small>{tax_now}</small>
<dialog id=taxdlg><form method=post action=/tax>
<b>Вид налогообложения</b><br><select name=mode id=taxmode onchange="taxdef()">{mode_opts}</select><br><br>
<b>Налоговая ставка, %</b><br>{radios}<br>
<input type=submit value=Сохранить> <button type=button onclick="document.getElementById('taxdlg').close()">Отмена</button>
</form></dialog>
<script>
var DEF={{'УСН «Доходы»':'6','УСН «Доходы минус расходы»':'15','НПД':'other','ОСН':'25'}};
function taxdef(){{var v=DEF[document.getElementById('taxmode').value];
var el=document.querySelector('input[name=rate][value="'+v+'"]');if(el)el.checked=true;}}
</script>"""

    return f"""<html><head><meta charset=utf-8><title>Финотчёт ВБ по артикулам</title>
<style>
body{{background:#1E1E1E;color:#EEEEEE;font:13px Verdana,sans-serif;margin:16px}}
a{{color:#EEEEEE}}
table{{border-collapse:collapse}}
td,th{{border:1px solid #343A52;padding:4px 8px;text-align:left}}
.up td{{background:#282828;vertical-align:top;width:33%}}
.up b{{display:block;margin-bottom:6px}}
.up input,.up label,.up button{{margin:3px 0}}
input,select,button{{background:#343A52;color:#EEEEEE;border:1px solid #2D4131;padding:3px 8px}}
input[type=submit]:hover,button:hover{{background:#2D4131;cursor:pointer}}
.d td,.d th{{font-size:clamp(9px,1.05vw,13px);padding:2px 6px;white-space:nowrap;background:#282828;text-align:right}}
.d th{{background:#343A52;text-align:center}}
.d td:first-child{{text-align:left}}
.d .tot td{{background:#2D4131;font-weight:bold}}
dialog{{background:#282828;color:#EEEEEE;border:1px solid #343A52;padding:16px}}
dialog::backdrop{{background:#1E1E1ECC}}
small{{color:#EEEEEE}}
</style></head>
<body><h2>Финотчёт ВБ — разрез по артикулам</h2>
<table class=up width=100%><tr>
<td><b>1. Детализация отчёта (xlsx)</b>
<form method=post action=/report enctype=multipart/form-data>
<input type=file name=f required> <input type=submit value=Загрузить></form>
<form method=post action=/skip>
<label><input type=checkbox name=on value=1 {"checked" if skip else ""} onchange="this.form.submit()">
Не учитывать «Оказание услуг WB Продвижение» в прочих удержаниях</label></form>
{modal}</td>
<td><b>2. Себестоимость</b> <small>(Артикул ВБ | Себестоимость)</small>
<form method=post action=/cost enctype=multipart/form-data>
<input type=file name=f required><br><input type=submit value=Загрузить></form></td>
<td><b>3. Реклама</b> <small>(ID РК | Артикул ВБ | Затраты)</small>
<form method=post action=/adv enctype=multipart/form-data>
<input type=file name=f required><br><input type=submit value=Загрузить></form></td>
</tr></table><br>
{f"<p>{export_link}</p>" if ART else ""}
<div style="overflow-x:auto"><table class=d>
<tr>{head_html}</tr>
{foot if ART else ""}{body}</table></div>
{"<p><i>Загрузите детализацию еженедельного отчёта.</i></p>" if not ART else ""}
<p><small>Прибыль = К перечислению − логистика − хранение − приёмка − штрафы − прочие удержания − себестоимость×(продажи−возвраты) − реклама − налог.
Налоговая база: УСН «Доходы»/НПД — «ВБ реализовал»; УСН «Д−Р»/ОСН — прибыль до налога (если &gt;0). Маржа = прибыль / «ВБ реализовал».</small></p></body></html>"""


def selftest():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Код номенклатуры", "Артикул поставщика", "Тип документа", "Обоснование для оплаты",
               "Кол-во", "Вайлдберриз реализовал товар (Пр)", "К перечислению Продавцу за реализованный товар",
               "Услуги по доставке товара покупателю", "Общая сумма штрафов", "Хранение",
               "Операции при приёмке", "Удержания", "Виды логистики, штрафов и корректировок ВВ"])
    ws.append([111, "ART-1", "Продажа", "Продажа", 1, 1000, 800, 0, 0, 0, 0, 0, ""])
    ws.append([111, "ART-1", "Возврат", "Возврат", 1, 500, 400, 0, 0, 0, 0, 0, ""])
    ws.append([111, "", "", "Логистика", 0, 0, 0, 70, 0, 0, 0, 0, ""])
    ws.append(["", "", "", "Хранение", 0, 0, 0, 0, 0, 30, 0, 0, ""])
    ws.append(["", "", "", "Удержания", 0, 0, 0, 0, 0, 0, 0, 250, "Подписка Джем"])
    ws.append(["", "", "", "Удержания", 0, 0, 0, 0, 0, 0, 0, 999, "Оказание услуг «WB Продвижение»"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    parse_report(buf)
    COST["111"] = 100
    ADV["111"] = 50
    a = ART["111"]
    assert a["pay"] == 400 and a["sold"] == 500 and a["logistics"] == 70
    assert ART["— без артикула"]["storage"] == 30
    assert OTHER["— без артикула"]["Подписка Джем"] == 250
    # прибыль 111 = 400 − 70 − 100×(1−1)=0 − 50 = 280
    html = app.test_client().get("/").text
    assert "280" in html and "ART-1" in html
    html = app.test_client().get("/?other=1").text
    assert "Подписка Джем" in html and "250" in html
    assert ART["— без артикула"]["other"] == 250 + 999
    c = app.test_client()
    c.post("/skip", data={"on": "1"})  # живой фильтр: данные не перегружаем
    html = c.get("/?other=1").text
    assert "услуг «WB" not in html and ">250<" in html  # столбца WB Продвижения нет, прочие = 250
    c.post("/skip", data={})
    assert "услуг «WB" in c.get("/?other=1").text  # выключили — вернулся
    c.post("/skip", data={"on": "1"})
    c.post("/tax", data={"mode": "УСН «Доходы»", "rate": "6", "custom": ""})
    assert tax_rate() == 6
    html = c.get("/?tax=1").text
    # артикул 111: база 500, налог 30, прибыль 280−30=250
    assert "Налоговая база" in html and ">30<" in html and ">250<" in html
    c.post("/tax", data={"mode": "УСН «Доходы»", "rate": "other", "custom": "6,5"})
    xls = load_workbook(io.BytesIO(c.get("/export?tax=1").data)).active
    flat = [v for row in xls.iter_rows(values_only=True) for v in row]
    assert "↳ Ставка, %" in flat and "6,5" in flat  # запятая в десятичных
    assert 32 in flat  # налог 111: 500×6.5% = 32.5 → 32 (банковское округление)
    c.post("/tax", data={"mode": "НПД", "rate": "other", "custom": "4"})
    assert tax_rate() == 4
    c.post("/tax", data={"mode": "ОСН", "rate": "none"})
    assert tax_rate() is None and "Налог <a" not in c.get("/").text  # столбец «Налог» скрыт
    print("selftest ok")


if __name__ == "__main__":
    if "--selftest" in sys.argv:
        selftest()
    else:
        app.run(port=8010)
